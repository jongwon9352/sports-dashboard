"""
전체 엑셀 데이터 → Supabase 마이그레이션 스크립트
3개 엑셀 파일의 모든 시트에서 데이터를 추출하여 Supabase에 삽입합니다.
"""
import openpyxl, json, ssl, urllib.request, sys, math
from datetime import datetime

ctx = ssl.create_default_context()

SUPABASE_URL = "https://ftockzbsceolvwztojpx.supabase.co"
ANON_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZ0b2NremJzY2VvbHZ3enRvanB4Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODIwMDY4NDgsImV4cCI6MjA5NzU4Mjg0OH0.o2oFkdg5IbzlhAhys9KLmtFQqI7txYMvNBf1Ep5Ntr4"

FILES = {
    'training': "/Users/apple/Desktop/2026 훈련 및 운동부하 데이터(custom data).xlsx",
    'match': "/Users/apple/Desktop/2026 경기 데이터 Vol3.xlsx",
    'physical': "/Users/apple/Desktop/2026 U15 피지컬 리포트.xlsx",
}

def num(v):
    if v is None: return None
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else round(f, 2)
    except: return None

def num0(v):
    r = num(v)
    return r if r is not None else 0

def date_str(v):
    if v and hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return None

def api_get(path):
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/{path}",
        headers={"apikey": ANON_KEY, "Authorization": f"Bearer {ANON_KEY}"}
    )
    with urllib.request.urlopen(req, context=ctx) as resp:
        return json.loads(resp.read())

def api_post(table, rows, batch_size=200):
    total = 0
    errors = 0
    for i in range(0, len(rows), batch_size):
        batch = rows[i:i+batch_size]
        body = json.dumps(batch).encode()
        req = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/{table}",
            data=body,
            headers={
                "apikey": ANON_KEY,
                "Authorization": f"Bearer {ANON_KEY}",
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates,return=minimal"
            },
            method="POST"
        )
        try:
            with urllib.request.urlopen(req, context=ctx) as resp:
                total += len(batch)
        except Exception as e:
            err = e.read().decode() if hasattr(e, 'read') else str(e)
            # Try one by one on batch failure
            for row in batch:
                body2 = json.dumps([row]).encode()
                req2 = urllib.request.Request(
                    f"{SUPABASE_URL}/rest/v1/{table}",
                    data=body2,
                    headers={
                        "apikey": ANON_KEY,
                        "Authorization": f"Bearer {ANON_KEY}",
                        "Content-Type": "application/json",
                        "Prefer": "resolution=merge-duplicates,return=minimal"
                    },
                    method="POST"
                )
                try:
                    with urllib.request.urlopen(req2, context=ctx) as resp2:
                        total += 1
                except:
                    errors += 1
    return total, errors

# Get player map
print("🔑 Getting player IDs...")
players = api_get("players?select=id,name")
PLAYER_MAP = {p['name']: p['id'] for p in players}
print(f"   {len(PLAYER_MAP)} players loaded")

def pid(name):
    if not name: return None
    return PLAYER_MAP.get(str(name).strip())

results = {}

# ============================================================
# 1. MATCH DATA (경기 데이터 Vol3 → Raw data sheet, 964행)
# ============================================================
print("\n" + "=" * 60)
print("📋 [1/6] 경기 데이터 Raw data (964행)")
print("=" * 60)

wb = openpyxl.load_workbook(FILES['match'], data_only=True, read_only=True)
ws = wb['Raw data']
match_rows = []
for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=25, values_only=True):
    opponent = row[0]
    dt = date_str(row[1])
    name = row[4]
    if not name or not dt: continue
    p = pid(name)
    if not p: continue

    match_rows.append({
        'player_id': p,
        'match_date': dt,
        'opponent': str(opponent) if opponent else None,
        'event_type': str(row[5]) if row[5] else None,
        'player_group': str(row[6]) if row[6] else None,
        'play_time_min': num0(row[7]),
        'rpe': num(row[8]),
        'total_distance': num0(row[9]),
        'speed_zone_1': num0(row[10]),
        'speed_zone_2': num0(row[11]),
        'speed_zone_3': num0(row[12]),
        'speed_zone_4': num0(row[13]),
        'speed_zone_5': num0(row[14]),
        'm_per_min': num0(row[15]),
        'hsr_distance': num0(row[16]),
        'sprint_distance': num0(row[17]),
        'sprint_count': num0(row[18]),
        'acc_count': num0(row[19]),
        'dec_count': num0(row[20]),
        'acd_load': num0(row[21]),
        'max_speed': num0(row[22]),
        'action_count': num0(row[23]) if len(row) > 23 and row[23] else num0(row[19]) + num0(row[20]),
    })
wb.close()

total, errs = api_post('match_data', match_rows)
results['match_data'] = {'parsed': len(match_rows), 'inserted': total, 'errors': errs}
print(f"   ✅ {total}/{len(match_rows)} rows inserted ({errs} errors)")

# ============================================================
# 2. MATURATION RPE (훈련 데이터 → Maturation RPE sheet, 162행)
# ============================================================
print("\n" + "=" * 60)
print("📋 [2/6] Maturation RPE (162행)")
print("=" * 60)

wb = openpyxl.load_workbook(FILES['training'], data_only=True, read_only=True)
ws = wb['Maturation RPE']
rpe_rows = []
for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=11, values_only=True):
    name = row[1]
    if not name: continue
    name = str(name).strip()
    dt = date_str(row[3])
    if not dt: continue

    rpe_rows.append({
        'player_id': pid(name),
        'name': name,
        'maturation': str(row[2]) if row[2] else None,
        'session_date': dt,
        'topic': str(row[5]) if row[5] else None,
        'session_name': str(row[6]) if row[6] else None,
        'duration_min': num(row[7]),
        'pitch_size': str(row[8]) if row[8] else None,
        'rpe': num(row[9]),
    })
wb.close()

total, errs = api_post('maturation_rpe', rpe_rows)
results['maturation_rpe'] = {'parsed': len(rpe_rows), 'inserted': total, 'errors': errs}
print(f"   ✅ {total}/{len(rpe_rows)} rows inserted ({errs} errors)")

# ============================================================
# 3. NATIONAL TEAM DATA (대표팀 차출 데이터, 12행)
# ============================================================
print("\n" + "=" * 60)
print("📋 [3/6] 대표팀 차출 데이터 (12행)")
print("=" * 60)

wb = openpyxl.load_workbook(FILES['training'], data_only=True, read_only=True)
ws = wb['대표팀 차출 데이터']
nt_rows = []
# Row 3 has player name (박승준)
player_name = None
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=11, values_only=True):
    if row[1] and isinstance(row[1], str) and not hasattr(row[1], 'strftime'):
        # Check if it's a name (not a date)
        try:
            datetime.strptime(str(row[1]), '%Y-%m-%d')
        except:
            player_name = str(row[1]).strip()
            continue

    dt = date_str(row[1])
    if not dt: continue

    nt_rows.append({
        'player_name': player_name or '',
        'callup_date': dt,
        'total_distance': num0(row[3]),
        'hsr_distance': num0(row[4]),
        'sprint_distance': num0(row[5]),
        'acc_count': num0(row[6]),
        'dec_count': num0(row[7]),
        'rpe': num0(row[8]),
        'notes': str(row[9]) if row[9] else None,
    })
wb.close()

total, errs = api_post('national_team_data', nt_rows)
results['national_team_data'] = {'parsed': len(nt_rows), 'inserted': total, 'errors': errs}
print(f"   ✅ {total}/{len(nt_rows)} rows inserted ({errs} errors)")

# ============================================================
# 4. PHYSICAL REPORT - Fix VALD + Sprint/Jump/COD
# ============================================================
print("\n" + "=" * 60)
print("📋 [4/6] Physical Report - VALD/Sprint/Jump 재이관")
print("=" * 60)

wb = openpyxl.load_workbook(FILES['physical'], data_only=True, read_only=False)
ws = wb['RAW Data']

# Read header row (row 6) to map column indices correctly
headers = {}
for col_idx, cell in enumerate(ws[6], start=0):
    if cell.value:
        headers[col_idx] = str(cell.value).strip()

print(f"   Found {len(headers)} header columns")

# Find VALD columns by header names
vald_cols = {}
for idx, name in headers.items():
    if '햄스트링 평균' in name: vald_cols['nordic_avg'] = idx
    elif '햄스트링 왼쪽' in name: vald_cols['nordic_left'] = idx
    elif '햄스트링 오른쪽' in name: vald_cols['nordic_right'] = idx
    elif '외전 평균' in name: vald_cols['hip_ab_avg'] = idx
    elif '외전 왼쪽' in name: vald_cols['hip_ab_left'] = idx
    elif '외전 오른쪽' in name: vald_cols['hip_ab_right'] = idx
    elif '내전 평균' in name: vald_cols['hip_ad_avg'] = idx
    elif '내전 왼쪽' in name: vald_cols['hip_ad_left'] = idx
    elif '내전 오른쪽' in name: vald_cols['hip_ad_right'] = idx
    elif '총 시간' in name and 'sprint' not in name.lower(): vald_cols['sprint_total'] = idx
    elif '5m 시간' in name: vald_cols['sprint_5m'] = idx
    elif '10m 시간' in name: vald_cols['sprint_10m'] = idx
    elif '30m 시간' in name: vald_cols['sprint_30m'] = idx
    elif '첫점프' in name: vald_cols['cmj'] = idx
    elif '재점프' in name: vald_cols['rebound'] = idx
    elif '스쿼트 점프' in name: vald_cols['squat_jump'] = idx
    elif '방향전환(런)' in name: vald_cols['cod_run'] = idx
    elif '방향전환(볼)' in name: vald_cols['cod_ball'] = idx

print(f"   VALD column mapping: {json.dumps({k: v for k, v in vald_cols.items()}, indent=2)}")

# Delete existing physical_report and re-insert with correct data
physical_rows = []
for row_idx in range(7, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=2).value  # B
    if not name or not str(name).strip(): continue
    name = str(name).strip()
    p = pid(name)
    if not p: continue

    test_date = ws.cell(row=row_idx, column=6).value  # F
    dt = date_str(test_date)
    if not dt: continue

    def cell_val(col_key):
        if col_key not in vald_cols: return None
        v = ws.cell(row=row_idx, column=vald_cols[col_key] + 1).value  # +1 for 1-based
        return num(v)

    r = {
        'player_id': p,
        'test_date': dt,
        'test_round': str(ws.cell(row=row_idx, column=5).value) if ws.cell(row=row_idx, column=5).value else '1',
        'height': num(ws.cell(row=row_idx, column=10).value),   # J
        'weight': num(ws.cell(row=row_idx, column=11).value),   # K
        'leg_length': num(ws.cell(row=row_idx, column=12).value), # L
        'sitting_height': num(ws.cell(row=row_idx, column=15).value), # O
        'maturity_offset': num(ws.cell(row=row_idx, column=43).value), # AQ
        'predicted_adult_height': num(ws.cell(row=row_idx, column=34).value), # AH
        'phv_age': num(ws.cell(row=row_idx, column=42).value), # AP
        'nordic_curl_left': cell_val('nordic_left'),
        'nordic_curl_right': cell_val('nordic_right'),
        'hip_ab_left': cell_val('hip_ab_left'),
        'hip_ab_right': cell_val('hip_ab_right'),
        'hip_ad_left': cell_val('hip_ad_left'),
        'hip_ad_right': cell_val('hip_ad_right'),
        'sprint_5m_time': cell_val('sprint_5m'),
        'sprint_10m_time': cell_val('sprint_10m'),
        'sprint_30m_time': cell_val('sprint_total'),
        'cmj_height': cell_val('cmj'),
        'rebound_jump_height': cell_val('rebound'),
        'squat_jump_height': cell_val('squat_jump'),
        'cod_run': cell_val('cod_run'),
        'cod_ball': cell_val('cod_ball'),
    }
    physical_rows.append(r)

wb.close()

# Delete existing and re-insert
req = urllib.request.Request(
    f"{SUPABASE_URL}/rest/v1/physical_report?id=not.is.null",
    headers={"apikey": ANON_KEY, "Authorization": f"Bearer {ANON_KEY}"},
    method="DELETE"
)
try:
    urllib.request.urlopen(req, context=ctx)
except: pass

total, errs = api_post('physical_report', physical_rows)
results['physical_report'] = {'parsed': len(physical_rows), 'inserted': total, 'errors': errs}
print(f"   ✅ {total}/{len(physical_rows)} rows inserted ({errs} errors)")

# ============================================================
# 5. GROWTH TRACKING (월별 키/체중 이력)
# ============================================================
print("\n" + "=" * 60)
print("📋 [5/6] Growth Tracking (월별 키/체중)")
print("=" * 60)

wb = openpyxl.load_workbook(FILES['physical'], data_only=True, read_only=False)
ws = wb['RAW Data']

# Columns 46-57 = 1월~12월 신장 (AT-BE)
# Columns 58-69 = 1월~12월 체중 (BF-BQ)
growth_rows = []
for row_idx in range(7, ws.max_row + 1):
    name = ws.cell(row=row_idx, column=2).value
    if not name or not str(name).strip(): continue
    name = str(name).strip()
    p = pid(name)
    if not p: continue

    for month in range(1, 13):
        h_col = 45 + month  # Columns 46-57 for height (1월~12월)
        w_col = 57 + month  # Columns 58-69 for weight (1월~12월)
        h = num(ws.cell(row=row_idx, column=h_col).value)
        w = num(ws.cell(row=row_idx, column=w_col).value)
        if h or w:
            growth_rows.append({
                'player_id': p,
                'year': 2026,
                'month': month,
                'height': h,
                'weight': w,
            })

wb.close()

total, errs = api_post('growth_tracking', growth_rows)
results['growth_tracking'] = {'parsed': len(growth_rows), 'inserted': total, 'errors': errs}
print(f"   ✅ {total}/{len(growth_rows)} rows inserted ({errs} errors)")

# ============================================================
# 6. VERIFY EXISTING DATA
# ============================================================
print("\n" + "=" * 60)
print("📋 [6/6] 기존 데이터 검증")
print("=" * 60)

tables = ['players', 'training_daily', 'acwr_daily', 'physical_report',
          'match_data', 'maturation_rpe', 'national_team_data', 'growth_tracking']

for tbl in tables:
    data = api_get(f"{tbl}?select=id&limit=1&order=id")
    # Get count via HEAD
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/{tbl}?select=id",
        headers={
            "apikey": ANON_KEY,
            "Authorization": f"Bearer {ANON_KEY}",
            "Prefer": "count=exact",
            "Range-Unit": "items",
            "Range": "0-0"
        }
    )
    try:
        with urllib.request.urlopen(req, context=ctx) as resp:
            content_range = resp.headers.get('Content-Range', '')
            count = content_range.split('/')[-1] if '/' in content_range else '?'
    except:
        count = '?'
    print(f"   {tbl}: {count} rows")

# ============================================================
# SUMMARY
# ============================================================
print("\n" + "=" * 60)
print("📊 MIGRATION SUMMARY")
print("=" * 60)
for tbl, r in results.items():
    status = "✅" if r['errors'] == 0 else "⚠️"
    print(f"   {status} {tbl}: {r['inserted']}/{r['parsed']} ({r['errors']} errors)")
print()
