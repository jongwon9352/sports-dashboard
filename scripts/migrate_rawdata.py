"""
RAWDATA.xlsx → Supabase training_daily 마이그레이션
CSV 업로드된 날짜는 건너뛴다.
"""
import openpyxl, datetime, uuid, unicodedata, math
from supabase import create_client

SUPABASE_URL = "https://ftockzbsceolvwztojpx.supabase.co"
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImZ0b2NremJzY2VvbHZ3enRvanB4Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODIwMDY4NDgsImV4cCI6MjA5NzU4Mjg0OH0.o2oFkdg5IbzlhAhys9KLmtFQqI7txYMvNBf1Ep5Ntr4"
XLSX_PATH = "/Users/apple/Desktop/RAWDATA.xlsx"

sb = create_client(SUPABASE_URL, SUPABASE_KEY)
DAY_NAMES = ['월', '화', '수', '목', '금', '토', '일']

def nfc(s):
    return unicodedata.normalize('NFC', str(s).strip())

def safe_float(v):
    if v is None: return 0
    try:
        f = float(v)
        return 0 if (math.isnan(f) or math.isinf(f)) else round(f, 2)
    except:
        return 0

def safe_nullable(v):
    if v is None or v == '' or v == '-': return None
    try:
        f = float(v)
        return None if (math.isnan(f) or math.isinf(f)) else round(f, 2)
    except:
        return None

def get_csv_uploaded_dates():
    dates = set()
    resp = sb.table('csv_uploads').select('training_date').execute()
    for row in resp.data or []:
        if row.get('training_date'):
            dates.add(row['training_date'])
    print(f"  CSV 업로드된 날짜: {len(dates)}개 → 건너뜀")
    return dates

def get_or_create_players(names_numbers):
    existing = {}
    resp = sb.table('players').select('id, name').execute()
    for p in resp.data or []:
        existing[nfc(p['name'])] = p['id']

    missing = []
    for name, num in names_numbers.items():
        if name not in existing:
            pid = str(uuid.uuid4())
            existing[name] = pid
            now = datetime.datetime.now(datetime.timezone.utc).isoformat()
            missing.append({
                'id': pid, 'name': name, 'jersey_number': num,
                'position': 'MF', 'grade': 'U15',
                'birth_date': None, 'maturity_status': 'Mid',
                'maturity_offset': 0, 'predicted_adult_height': 0,
                'current_height': 0, 'current_weight': 0,
                'latest_mas': None, 'latest_mss': None,
                'created_at': now, 'updated_at': now,
            })

    if missing:
        for i in range(0, len(missing), 50):
            sb.table('players').insert(missing[i:i+50]).execute()
        print(f"  새 선수 {len(missing)}명 생성")

    return existing

def read_xlsx():
    print(f"📂 엑셀 파일 읽는 중: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = []
    for row in ws.iter_rows(min_row=8, max_col=28, values_only=True):
        date_val = row[1]
        name_val = row[4]
        if not isinstance(date_val, datetime.datetime) or not name_val:
            continue
        rows.append({
            'date': date_val.strftime('%Y-%m-%d'),
            'day_of_week': str(row[2] or ''),
            'week_label': str(row[3] or ''),
            'name': nfc(name_val),
            'jersey_number': int(row[26]) if row[26] and str(row[26]).isdigit() else 0,
            'duration_min': safe_float(row[7]),
            'rpe': safe_nullable(row[8]),
            'total_distance': safe_float(row[9]),
            'speed_zone_1': safe_float(row[10]),
            'speed_zone_2': safe_float(row[11]),
            'speed_zone_3': safe_float(row[12]),
            'speed_zone_4': safe_float(row[13]),
            'speed_zone_5': safe_float(row[14]),
            'm_per_min': safe_float(row[15]),
            'hsr_distance': safe_float(row[16]),
            'sprint_distance': safe_float(row[17]),
            'sprint_count': safe_float(row[18]),
            'acc_count': safe_float(row[19]),
            'dec_count': safe_float(row[20]),
            'acd_load': safe_float(row[21]),
            'max_speed': safe_float(row[22]),
        })
    wb.close()
    print(f"  총 {len(rows)}행 읽음")
    return rows

def migrate():
    print("🚀 마이그레이션 시작\n")

    skip_dates = get_csv_uploaded_dates()
    raw_rows = read_xlsx()

    # 건너뛸 날짜 제외
    rows = [r for r in raw_rows if r['date'] not in skip_dates]
    skipped = len(raw_rows) - len(rows)
    print(f"  CSV 업로드 날짜 제외: {skipped}행 건너뜀 → {len(rows)}행 대상\n")

    if not rows:
        print("✅ 마이그레이션할 데이터가 없습니다.")
        return

    # 선수 생성/조회
    names_numbers = {}
    for r in rows:
        if r['name'] not in names_numbers:
            names_numbers[r['name']] = r['jersey_number']
    player_map = get_or_create_players(names_numbers)

    # 같은 선수+날짜가 여러 세션이면 합산 (max_speed는 max)
    now = datetime.datetime.now(datetime.timezone.utc).isoformat()
    SUM_KEYS = ['duration_min', 'total_distance', 'speed_zone_1', 'speed_zone_2',
                'speed_zone_3', 'speed_zone_4', 'speed_zone_5', 'hsr_distance',
                'sprint_distance', 'sprint_count', 'acc_count', 'dec_count', 'acd_load']

    grouped = {}
    for r in rows:
        pid = player_map.get(r['name'])
        if not pid:
            continue
        key = (pid, r['date'])
        if key not in grouped:
            grouped[key] = {
                'player_id': pid,
                'training_date': r['date'],
                'day_of_week': r['day_of_week'],
                'week_label': r['week_label'],
                'rpe': r['rpe'],
                'max_speed': r['max_speed'],
                'm_per_min': r['m_per_min'],
            }
            for k in SUM_KEYS:
                grouped[key][k] = r[k]
        else:
            g = grouped[key]
            for k in SUM_KEYS:
                g[k] = round(g[k] + r[k], 2)
            g['max_speed'] = max(g['max_speed'], r['max_speed'])
            if r['rpe'] is not None:
                g['rpe'] = r['rpe']

    db_rows = []
    for g in grouped.values():
        duration = g['duration_min']
        td = g['total_distance']
        rpe = g['rpe']
        daily_tl = round(rpe * duration, 2) if rpe is not None else None
        m_per_min = round(td / duration, 1) if duration > 0 else 0

        db_rows.append({
            'id': str(uuid.uuid4()),
            'player_id': g['player_id'],
            'training_date': g['training_date'],
            'day_of_week': g['day_of_week'],
            'week_label': g['week_label'],
            'duration_min': duration,
            'rpe': rpe,
            'total_distance': td,
            'm_per_min': m_per_min,
            'speed_zone_1': g['speed_zone_1'],
            'speed_zone_2': g['speed_zone_2'],
            'speed_zone_3': g['speed_zone_3'],
            'speed_zone_4': g['speed_zone_4'],
            'speed_zone_5': g['speed_zone_5'],
            'hsr_distance': g['hsr_distance'],
            'hsr_custom': 0,
            'sprint_distance': g['sprint_distance'],
            'sprint_custom': 0,
            'sprint_count': g['sprint_count'],
            'sprint_count_custom': 0,
            'acc_count': g['acc_count'],
            'dec_count': g['dec_count'],
            'acd_load': g['acd_load'],
            'max_speed': g['max_speed'],
            'daily_training_load': daily_tl,
            'created_at': now,
        })

    print(f"  DB 저장 대상: {len(db_rows)}행")

    # upsert in batches
    batch_size = 100
    for i in range(0, len(db_rows), batch_size):
        batch = db_rows[i:i+batch_size]
        sb.table('training_daily').upsert(batch, on_conflict='player_id,training_date').execute()
        print(f"  ✓ {min(i+batch_size, len(db_rows))}/{len(db_rows)} 저장 완료")

    # 고유 날짜 목록
    dates = sorted(set(r['date'] for r in rows))
    print(f"\n✅ 마이그레이션 완료!")
    print(f"   {len(db_rows)}행 · {len(dates)}일 · {len(player_map)}명")
    print(f"   날짜 범위: {dates[0]} ~ {dates[-1]}")

if __name__ == '__main__':
    migrate()
